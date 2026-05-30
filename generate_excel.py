from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()

# 创建样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(bold=True, size=14)
center_align = Alignment(horizontal="center")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# ==================== 运动功能数据统计 ====================
ws1 = wb.active
ws1.title = "运动功能数据统计"

# 标题
ws1.merge_cells('A1:D1')
ws1['A1'] = "运动功能数据统计"
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

# 运动记录表
row = 3
ws1[f'A{row}'] = "一、运动记录 (WorkoutRecord)"
ws1[f'A{row}'].font = Font(bold=True, size=11)
row += 1

headers = ["字段名", "数据类型", "默认值", "说明"]
for col, header in enumerate(headers, 1):
    cell = ws1[f'{get_column_letter(col)}{row}']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

rows = [
    ["id", "Int", "-", "主键ID"],
    ["userId", "Int", "-", "用户ID"],
    ["type", "String", "-", "运动类型（跑步、游泳、健身、骑行等）"],
    ["durationMinutes", "Int", "-", "运动时长（分钟）"],
    ["calories", "Int", "0", "消耗卡路里"],
    ["recordDate", "DateTime", "-", "记录日期"],
    ["notes", "String", "-", "备注说明"],
    ["createdAt", "DateTime", "now()", "创建时间"],
]

for r in rows:
    for col, val in enumerate(r, 1):
        cell = ws1[f'{get_column_letter(col)}{row}']
        cell.value = val
        cell.border = thin_border
        cell.alignment = center_align if col <= 3 else Alignment(horizontal="left")
    row += 1

# 运动目标表
row += 2
ws1[f'A{row}'] = "二、运动目标 (Goal)"
ws1[f'A{row}'].font = Font(bold=True, size=11)
row += 1

for col, header in enumerate(headers, 1):
    cell = ws1[f'{get_column_letter(col)}{row}']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

rows = [
    ["id", "Int", "-", "主键ID"],
    ["userId", "Int", "-", "用户ID"],
    ["goalType", "GoalType", "-", "目标类型（每日运动分钟数/睡眠小时数/饮食卡路里）"],
    ["targetValue", "Int", "-", "目标数值"],
    ["period", "GoalPeriod", "DAILY", "周期（每日）"],
    ["createdAt", "DateTime", "now()", "创建时间"],
    ["updatedAt", "DateTime", "-", "更新时间"],
]

for r in rows:
    for col, val in enumerate(r, 1):
        cell = ws1[f'{get_column_letter(col)}{row}']
        cell.value = val
        cell.border = thin_border
        cell.alignment = center_align if col <= 3 else Alignment(horizontal="left")
    row += 1

# 调整列宽
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 45

# ==================== 工作功能数据统计 ====================
ws2 = wb.create_sheet(title="工作功能数据统计")

# 标题
ws2.merge_cells('A1:D1')
ws2['A1'] = "工作功能数据统计"
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

# 工作设置表
row = 3
ws2[f'A{row}'] = "一、工作设置 (WorkSettings)"
ws2[f'A{row}'].font = Font(bold=True, size=11)
row += 1

for col, header in enumerate(headers, 1):
    cell = ws2[f'{get_column_letter(col)}{row}']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

rows = [
    ["id", "Int", "-", "主键ID"],
    ["userId", "Int", "-", "用户ID（唯一约束）"],
    ["occupation", "String", "-", "职业类型"],
    ["pomodoroDuration", "Int", "25", "番茄钟时长（分钟）"],
    ["sedentaryReminderOn", "Boolean", "true", "久坐提醒开关"],
    ["sedentaryInterval", "Int", "60", "久坐提醒间隔（分钟）"],
    ["wristHealthScore", "Int", "0", "手腕健康分数（程序员）"],
    ["eyeRestCount", "Int", "0", "眼睛休息次数"],
    ["waterIntake", "Int", "0", "饮水量（杯）（教师）"],
    ["backRelaxCount", "Int", "0", "背部放松次数（司机）"],
    ["vocalRestCount", "Int", "0", "嗓音休息次数（教师）"],
    ["eyeExerciseCount", "Int", "0", "眼操次数（学生）"],
    ["deepBreathCount", "Int", "0", "深呼吸次数（医护）"],
    ["neckRelaxCount", "Int", "0", "肩颈放松次数（办公）"],
    ["stepCount", "Int", "0", "步数（外勤销售）"],
    ["standCount", "Int", "0", "站立次数（通用）"],
    ["createdAt", "DateTime", "now()", "创建时间"],
    ["updatedAt", "DateTime", "-", "更新时间"],
]

for r in rows:
    for col, val in enumerate(r, 1):
        cell = ws2[f'{get_column_letter(col)}{row}']
        cell.value = val
        cell.border = thin_border
        cell.alignment = center_align if col <= 3 else Alignment(horizontal="left")
    row += 1

# 工作会话表
row += 2
ws2[f'A{row}'] = "二、工作会话 (WorkSession)"
ws2[f'A{row}'].font = Font(bold=True, size=11)
row += 1

for col, header in enumerate(headers, 1):
    cell = ws2[f'{get_column_letter(col)}{row}']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

rows = [
    ["id", "Int", "-", "主键ID"],
    ["userId", "Int", "-", "用户ID"],
    ["type", "String", "pomodoro", "会话类型（番茄钟/休息等）"],
    ["startTime", "DateTime", "-", "开始时间"],
    ["endTime", "DateTime", "-", "结束时间"],
    ["duration", "Int", "0", "持续时长（分钟）"],
    ["createdAt", "DateTime", "now()", "创建时间"],
]

for r in rows:
    for col, val in enumerate(r, 1):
        cell = ws2[f'{get_column_letter(col)}{row}']
        cell.value = val
        cell.border = thin_border
        cell.alignment = center_align if col <= 3 else Alignment(horizontal="left")
    row += 1

# 工作待办表
row += 2
ws2[f'A{row}'] = "三、工作待办 (WorkTodo)"
ws2[f'A{row}'].font = Font(bold=True, size=11)
row += 1

for col, header in enumerate(headers, 1):
    cell = ws2[f'{get_column_letter(col)}{row}']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

rows = [
    ["id", "Int", "-", "主键ID"],
    ["userId", "Int", "-", "用户ID"],
    ["content", "String", "-", "待办内容"],
    ["completed", "Boolean", "false", "是否完成"],
    ["todoDate", "DateTime", "-", "待办日期"],
    ["createdAt", "DateTime", "now()", "创建时间"],
]

for r in rows:
    for col, val in enumerate(r, 1):
        cell = ws2[f'{get_column_letter(col)}{row}']
        cell.value = val
        cell.border = thin_border
        cell.alignment = center_align if col <= 3 else Alignment(horizontal="left")
    row += 1

# 久坐响应表
row += 2
ws2[f'A{row}'] = "四、久坐响应 (SedentaryResponse)"
ws2[f'A{row}'].font = Font(bold=True, size=11)
row += 1

for col, header in enumerate(headers, 1):
    cell = ws2[f'{get_column_letter(col)}{row}']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

rows = [
    ["id", "Int", "-", "主键ID"],
    ["userId", "Int", "-", "用户ID"],
    ["respondedAt", "DateTime", "now()", "响应时间"],
]

for r in rows:
    for col, val in enumerate(r, 1):
        cell = ws2[f'{get_column_letter(col)}{row}']
        cell.value = val
        cell.border = thin_border
        cell.alignment = center_align if col <= 3 else Alignment(horizontal="left")
    row += 1

# 调整列宽
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 45

# ==================== 汇总统计指标 ====================
ws3 = wb.create_sheet(title="汇总统计指标")

# 标题
ws3.merge_cells('A1:C1')
ws3['A1'] = "汇总统计指标"
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

# 表头
row = 3
headers = ["统计指标", "计算方式", "用途"]
for col, header in enumerate(headers, 1):
    cell = ws3[f'{get_column_letter(col)}{row}']
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
row += 1

rows = [
    ["今日运动时长", "当天所有运动记录时长之和", "展示当日运动量"],
    ["今日消耗卡路里", "当天所有运动记录卡路里之和", "热量消耗追踪"],
    ["目标完成率", "实际完成值 / 目标值 × 100%", "目标进度展示"],
    ["今日专注时长", "当天所有工作会话时长之和", "工作效率分析"],
    ["番茄钟完成数", "当天完成的番茄钟会话数", "专注度统计"],
    ["待办完成率", "已完成待办数 / 总待办数 × 100%", "任务管理"],
]

for r in rows:
    for col, val in enumerate(r, 1):
        cell = ws3[f'{get_column_letter(col)}{row}']
        cell.value = val
        cell.border = thin_border
        cell.alignment = center_align if col == 1 else Alignment(horizontal="left")
    row += 1

# 调整列宽
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 25

# 保存文件
wb.save("健康应用数据统计.xlsx")
print("Excel 文件已生成：健康应用数据统计.xlsx")